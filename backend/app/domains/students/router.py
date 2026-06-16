from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import select as sa_select
from typing import Optional
import io
import openpyxl
from app.database import get_db
from app.core.auth import get_current_user
from app.domains.users.models import User
from app.domains.students.models import Program
from app.domains.students.schemas import (
    ProgramCreate,
    ProgramResponse,
    StudentCreate,
    StudentResponse,
    EnrollmentCreate,
    EnrollmentResponse,
)
from app.domains.students.services import (
    get_programs,
    create_program,
    update_program,
    delete_program,
    get_students,
    get_student,
    create_student,
    update_student,
    delete_student,
    get_enrollments,
    create_enrollment,
    delete_enrollment,
)

router = APIRouter(tags=["Students"])


# --- Programs ---
@router.get("/programs/", response_model=list[ProgramResponse])
def list_programs(
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    return get_programs(db, current_user.institution_id)


@router.post("/programs/", response_model=ProgramResponse, status_code=201)
def add_program(
    data: ProgramCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return create_program(db, data, current_user.institution_id)


@router.put("/programs/{program_id}", response_model=ProgramResponse)
def edit_program(
    program_id: str,
    data: ProgramCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return update_program(db, program_id, data, current_user.institution_id)


@router.delete("/programs/{program_id}", status_code=204)
def remove_program(
    program_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    delete_program(db, program_id, current_user.institution_id)


@router.get("/programs/export/xlsx")
def export_programs_xlsx(
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    programs = get_programs(db, current_user.institution_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Programas"
    ws.append(["Nombre", "Tipo certificado", "Total horas", "Resolución"])
    for p in programs:
        ws.append(
            [p.name, p.certificate_type or "", p.total_hours or "", p.resolution or ""]
        )
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=programas.xlsx"},
    )


@router.get("/programs/template/xlsx")
def download_programs_template(current_user: User = Depends(get_current_user)):
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Programas"
    headers = [
        "Nombre del programa",
        "Tipo de certificado",
        "Total horas",
        "N° Resolución",
    ]
    ws.append(headers)
    ws.append(
        ["Auxiliar de Enfermería", "Técnico Laboral en Salud", "1440", "001 de 2024"]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a2b4a")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_programas.xlsx"
        },
    )


@router.post("/programs/import/xlsx", status_code=201)
async def import_programs_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    created = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row[0]:
            continue
        try:
            data = ProgramCreate(
                name=str(row[0]).strip(),
                certificate_type=str(row[1]).strip() if row[1] else None,
                total_hours=str(row[2]).strip() if row[2] else None,
                resolution=str(row[3]).strip() if row[3] else None,
            )
            create_program(db, data, current_user.institution_id)
            created += 1
        except Exception as e:
            errors.append(f"Fila {i}: {str(e)}")
    return {
        "created": created,
        "errors": errors,
        "message": f"{created} programa(s) importado(s) exitosamente.",
    }


# --- Students ---
@router.get("/students/", response_model=list[StudentResponse])
def list_students(
    program_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return get_students(db, current_user.institution_id, program_id)


@router.get("/students/export/xlsx")
def export_students_xlsx(
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    students = get_students(db, current_user.institution_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"
    ws.append(
        [
            "Nombre completo",
            "Tipo documento",
            "N° Documento",
            "Lugar expedición",
            "Dirección",
            "Barrio",
            "Comuna",
            "Teléfono",
            "Email",
            "Menor de edad",
        ]
    )
    for s in students:
        ws.append(
            [
                s.full_name,
                s.document_type,
                s.document_number,
                s.document_place or "",
                s.address or "",
                s.neighborhood or "",
                s.commune or "",
                s.phone or "",
                s.email or "",
                "Sí" if s.is_minor else "No",
            ]
        )
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=estudiantes.xlsx"},
    )


@router.get("/students/template/xlsx")
def download_students_template(current_user: User = Depends(get_current_user)):
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"
    headers = [
        "Nombre completo",
        "Tipo documento (CC/TI/CE/PA/RC)",
        "N° Documento",
        "Lugar expedición",
        "Dirección",
        "Barrio",
        "Comuna",
        "Teléfono",
        "Email",
        "Nombre del programa",
        "Tipo de certificado",
        "Total horas",
        "N° Resolución",
        "Menor de edad (Si/No)",
        "Nombre representante",
        "Documento representante",
        "Dirección representante",
        "Teléfono representante",
    ]
    ws.append(headers)
    ws.append(
        [
            "María García López",
            "CC",
            "1234567890",
            "Medellín",
            "Calle 45 # 12-34",
            "Laureles",
            "11",
            "3001234567",
            "maria@ejemplo.com",
            "Auxiliar de Enfermería",
            "Técnico Laboral en Salud",
            "1440",
            "001 de 2024",
            "No",
            "",
            "",
            "",
            "",
        ]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a2b4a")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_estudiantes.xlsx"
        },
    )


@router.post("/students/import/xlsx", status_code=201)
async def import_students_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    created = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row[0]:
            continue
        try:
            # Columnas 9-12: datos del programa
            program_name = str(row[9]).strip() if len(row) > 9 and row[9] else None
            certificate_type = (
                str(row[10]).strip() if len(row) > 10 and row[10] else None
            )
            total_hours = str(row[11]).strip() if len(row) > 11 and row[11] else None
            resolution = str(row[12]).strip() if len(row) > 12 and row[12] else None

            print(f"Fila {i}: estudiante='{row[0]}' programa='{program_name}'")

            # Buscar o crear programa
            program = None
            if program_name:
                program = db.execute(
                    sa_select(Program).where(
                        Program.institution_id == current_user.institution_id,
                        Program.name == program_name,
                        Program.is_active == True,
                    )
                ).scalar_one_or_none()

                if not program:
                    prog_data = ProgramCreate(
                        name=program_name,
                        certificate_type=certificate_type,
                        total_hours=total_hours,
                        resolution=resolution,
                    )
                    program = create_program(db, prog_data, current_user.institution_id)
                    print(f"Programa creado: {program.name}")

            # Crear estudiante
            is_minor_val = (
                str(row[13]).strip().lower() in ["si", "sí", "yes", "true"]
                if len(row) > 13 and row[13]
                else False
            )
            data = StudentCreate(
                full_name=str(row[0]).strip(),
                document_type=str(row[1]).strip() if row[1] else "CC",
                document_number=str(row[2]).strip() if row[2] else "",
                document_place=str(row[3]).strip() if row[3] else None,
                address=str(row[4]).strip() if row[4] else None,
                neighborhood=str(row[5]).strip() if row[5] else None,
                commune=str(row[6]).strip() if row[6] else None,
                phone=str(row[7]).strip() if row[7] else None,
                email=str(row[8]).strip() if row[8] else None,
                is_minor=is_minor_val,
                guardian_name=(
                    str(row[14]).strip() if len(row) > 14 and row[14] else None
                ),
                guardian_document=(
                    str(row[15]).strip() if len(row) > 15 and row[15] else None
                ),
                guardian_address=(
                    str(row[16]).strip() if len(row) > 16 and row[16] else None
                ),
                guardian_phone=(
                    str(row[17]).strip() if len(row) > 17 and row[17] else None
                ),
            )
            student = create_student(db, data, current_user.institution_id)
            print(f"Estudiante creado: {student.full_name}")

            # Crear matrícula si hay programa
            if program:
                enrollment_data = EnrollmentCreate(
                    student_id=student.id,
                    program_id=program.id,
                )
                create_enrollment(db, enrollment_data, current_user.institution_id)
                print(f"Matrícula creada: {student.full_name} -> {program.name}")

            created += 1
        except Exception as e:
            print(f"ERROR fila {i}: {type(e).__name__}: {e}")
            errors.append(f"Fila {i}: {str(e)}")

    return {
        "created": created,
        "errors": errors,
        "message": f"{created} estudiante(s) importado(s) exitosamente.",
    }


@router.get("/students/{student_id}", response_model=StudentResponse)
def get_student_by_id(
    student_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return get_student(db, student_id, current_user.institution_id)


@router.post("/students/", response_model=StudentResponse, status_code=201)
def add_student(
    data: StudentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return create_student(db, data, current_user.institution_id)


@router.put("/students/{student_id}", response_model=StudentResponse)
def edit_student(
    student_id: str,
    data: StudentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return update_student(db, student_id, data, current_user.institution_id)


@router.delete("/students/{student_id}", status_code=204)
def remove_student(
    student_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    delete_student(db, student_id, current_user.institution_id)


# --- Enrollments ---
@router.get("/enrollments/", response_model=list[EnrollmentResponse])
def list_enrollments(
    program_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return get_enrollments(db, current_user.institution_id, program_id)


@router.get("/enrollments/export/xlsx")
def export_enrollments_xlsx(
    program_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    enrollments = get_enrollments(db, current_user.institution_id, program_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matrículas"
    ws.append(
        [
            "N° Matrícula",
            "Folio",
            "Nombre completo",
            "Tipo documento",
            "N° Documento",
            "Lugar expedición",
            "Dirección",
            "Barrio",
            "Comuna",
            "Teléfono",
            "Email",
            "Programa",
            "Tipo certificado",
            "Año",
        ]
    )
    for e in enrollments:
        ws.append(
            [
                e.enrollment_number or "",
                e.folio or "",
                e.student.full_name,
                e.student.document_type,
                e.student.document_number,
                e.student.document_place or "",
                e.student.address or "",
                e.student.neighborhood or "",
                e.student.commune or "",
                e.student.phone or "",
                e.student.email or "",
                e.program.name,
                e.certificate_type or "",
                e.year or "",
            ]
        )
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"matriculas_{program_id or 'todos'}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/enrollments/", response_model=EnrollmentResponse, status_code=201)
def add_enrollment(
    data: EnrollmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return create_enrollment(db, data, current_user.institution_id)


@router.delete("/enrollments/{enrollment_id}", status_code=204)
def remove_enrollment(
    enrollment_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    delete_enrollment(db, enrollment_id, current_user.institution_id)
